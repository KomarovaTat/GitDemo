import openpyxl


class HomePageData:
    data = [
        {"firstName": "Tatiana123", "email": "tatiana@gmail.com", "password": "12345", "checkbox": "true", "gender": 1},
        {"firstName": "John", "email": "john.doe@example.com", "password": "password123", "checkbox": "false",
         "gender": 0},
        {"firstName": "Alice", "email": "alice.smith@example.com", "password": "password456", "checkbox": "true",
         "gender": 1}]

    @staticmethod  # for static method we don't need to add #self
    def getTestData(test_case_name):
        result = {}
        book = openpyxl.load_workbook(
            "C:/Users/tkomarova/Desktop/Useful/Udemy/PythonData2.xlsx")  # add path of the PythonDemo file
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    result[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [result]
# {"firstName": "John", "email": "john.doe@example.com", "password": "password123", "checkbox": "false",
# then in the test_HomePage.py change 97 row from @pytest.fixture(params=HomePageData.data) to @pytest.fixture(params=HomePageData.getTestDate("Testcase2"))
